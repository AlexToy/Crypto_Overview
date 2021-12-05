from Currency import Currency
from Transaction import Transactions
import openpyxl
import re

DATE = 8
SIDE = 5
EXECUTED = 9 # BUY
TRADING = 11 # SELL
STATUS = 12
ORDER_NUM = 2

transactions_list = []

def reading_input_file_and_return_transactions_list(input_file):
    input_file = openpyxl.load_workbook(input_file, read_only=True)
    sheet = input_file.active
    transaction_list = []
    row_nbr = 1

    for row in sheet:
        row_nbr = row_nbr + 1
        if str(sheet.cell(row_nbr, STATUS).value) == "FILLED":
            side = str(sheet.cell(row_nbr, SIDE).value)
            date = str(sheet.cell(row_nbr, DATE).value)
            buy = str(sheet.cell(row_nbr, EXECUTED).value)
            order_num = str(sheet.cell(row_nbr, ORDER_NUM).value)
            quantity_buy = float(re.sub(",|[^\d|.]", "", buy))  # Delete the character ',' and leters
            currency_buy = re.sub("[^A-Z]", "", buy)  # Delete everthing except leters
            sell = str(sheet.cell(row_nbr, TRADING).value)
            quantity_sell = float(re.sub(",|[^\d|.]", "", sell))  # Delete the character ',' and leters
            currency_sell = re.sub("[^A-Z]", "", sell)  # Delete everthing except leters
            transaction_list.append(Transactions(currency_buy, quantity_buy, currency_sell, quantity_sell, side, date, order_num))
            #print("Trading appened : ", side, " ", quantityBuy, " ", currencyBuy, " for ", quantitySell, " ",
            #      currencySell)
        else:
            # print("Trading canceled")
            pass

    # sort by date the list of transactions
    transaction_list_sorted = sorted(transaction_list, key=lambda transaction: transaction.date)

    return transaction_list_sorted


def creation_of_trade_currency_list(transactions_list):
    currency_list = []
    for transaction in transactions_list:
        # Creation of a list with all currency
        currency_in_the_list = False  # reset each time there is a new transaction
        for currency in currency_list:
            if transaction.currencyBuy == currency.name:
                currency_in_the_list = True

        if currency_in_the_list is False or len(currency_list) == 0:
            currency_list.append(Currency(transaction.currencyBuy))

    return currency_list


def add_transaction_in_its_currency(transaction_list, currency_list):
    # Adds each transaction to its currency class
    for transaction in transaction_list:
        # print("Transaction currency : ", transaction.currencyBuy)
        for currency in currency_list:
            # print("Currency in the list : ", currency.name)
            if transaction.currencyBuy == currency.name:
                currency.add_transaction(transaction)
                """print("Trading add in ", currency.name, " : ", transaction.side, " ", transaction.quantityBuy, " ", transaction.currencyBuy, " for ", transaction.quantitySell, " ",
                      transaction.currencySell)"""

def add_new_transaction_manually():
    print("Add new transaction :")
    print("Currency buy :")
    currency_buy = input()
    print("Quantity buy :")
    quantity_buy = float(input())
    print("Currency sell :")
    currency_sell = input()
    print("Quantity sell :")
    quantity_sell = float(input())
    print("Side (BUY or SELL) :")
    side = input()
    print("Date (year-month-day xx:xx:xx) :")
    date = input()
    print("Number order:")
    order_num = input()
    return Transactions(currency_buy, quantity_buy, currency_sell, quantity_sell, side, date, order_num)


def option0(adding_type):
    # Add transaction list from an input file"
    if adding_type == "0":
        list_from_input_file = reading_input_file_and_return_transactions_list("Binance1116.xlsx")
        for transaction in list_from_input_file:
            transactions_list.append(transaction)
    # Add transaction manually
    elif adding_type == "1":
        transactions_list.append(add_new_transaction_manually())

    for t in transactions_list:
        print(t.transactionNumber)
    # sort by date the list of transactions0
    sorted(transactions_list, key=lambda transaction: transaction.date)
    # Create the list of currencies
    currencies_list = creation_of_trade_currency_list(transactions_list)
    # Add each transaction to his currency
    add_transaction_in_its_currency(transactions_list, currencies_list)
    # Calculation of transactions for all currencies
    for currency in currencies_list:
        currency.transactions_overview()

    return currencies_list

def option1(currencies_list):
    print(" List of currencies: ")
    for currency in currencies_list:
        print(currency.name)


def option2(currency_choose, currencies_list):
    for currency in currencies_list:
        if currency.name == currency_choose:
            currency.print_overview()


def main():

    while True:
        print("CHOOSE AN OPTION ")
        print("0 : Add transaction table")
        print("1 : Print list of currencies")
        print("2 : Currency overview")
        x = input()
        if x == "0":
            print("0 : Add transaction list from an input file")
            print("1 : Add transaction manually")
            adding_type = input()
            currencies_list = option0(adding_type)

        elif x == "1":
            option1(currencies_list)

        elif x == "2":
            print("What currency would you like to see ?")
            currency_choose = input()
            option2(currency_choose, currencies_list)


main()